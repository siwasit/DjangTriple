import os
from openpyxl import load_workbook
from django.shortcuts import redirect, render
from django.http import FileResponse, JsonResponse, HttpResponse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.urls import reverse

import rdflib
from rdflib import Graph, Literal, Namespace, RDF, XSD
import rdflib
from io import BytesIO
import plotly.graph_objs as go
import networkx as nx

from .utils import parse_date, graph_construct

from .models import ExcelFile
from .forms import addTriple , RegisterForm, UploadFileForm

def sign_in(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(
            request, 
            username=username, 
            password=password
        )

        if user is not None:
            # Log user in
            login(request, user)
            return redirect('homepage')
            
    return render(request, 'login.html')

def sign_up(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('homepage')
    else:
        form = RegisterForm()
        print(form)

    return render(request, 'register.html', {'form': form})

def sign_out(request):
    # sign user out
    logout(request)

    # Redirect to sign-in page
    return redirect('/sign-in')

@login_required(login_url="/sign-in")
def homepage(request, file_number=1):
    # for row in sheet.iter_rows(values_only=True):
    #     print(row)
    file_number = int(file_number) - 1
        
        # Fetch the ExcelFile object based on the provided file_number
    excel_file = ExcelFile.objects.all()[file_number]
    
    # Load the Excel workbook
    workbook = load_workbook(excel_file.file.path)
    print(excel_file.file.path)
    
    # Get the active sheet
    sheet = workbook.active

    items = []
    row_count = 1
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i==1:
            continue
        row_list = list(row)
        row_list.insert(0, (str(i-1)))
        items.append(row_list)
        row_count = i

    if request.method == "POST":
        add_form = addTriple(request.POST)
        if add_form.is_valid():
            triple_add = add_form.cleaned_data

            sheet["A" + str(row_count + 1)] = triple_add['subject']
            sheet["B" + str(row_count + 1)] = triple_add['predicate']
            sheet["C" + str(row_count + 1)] = triple_add['object']
            # print(type(triple_add['object']))

            workbook.save(excel_file.file.path)

            return redirect(reverse('homepage', kwargs={'file_number': file_number + 1}))
    else:
        add_form = addTriple()

    context = {
        'items': items,
        'add_form': add_form,
        'sheet_name': excel_file.name,
        'sheet_num': file_number,
    }

    return render(request, 'index.html', context)

def triple_delete(request, triple_id, sheet_num):
    excel_file = ExcelFile.objects.all()[sheet_num]
    excel_file_path = excel_file.file.path
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    sheet.delete_rows(idx=triple_id + 1)
    workbook.save(excel_file_path)
    return redirect(reverse('homepage', kwargs={'file_number': sheet_num + 1}))

def triple_edit(request, triple_id, sheet_num):
    excel_file = ExcelFile.objects.all()[sheet_num]
    excel_file_path = excel_file.file.path
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    if request.method == "POST":
        triple_form_get = request.POST
        sheet["A" + str(triple_id + 1)] = triple_form_get['editSubject']
        sheet["B" + str(triple_id + 1)] = triple_form_get['editPredicate']
        sheet["C" + str(triple_id + 1)] = triple_form_get['editObject']
        workbook.save(excel_file_path)
        return redirect(reverse('homepage', kwargs={'file_number': sheet_num + 1}))
    else:
        return JsonResponse({'Error': '404 internal server error'})
    
# @login_required(login_url="/sign-in")
def rdffile_export(request, sheet_num):

    rdf_data = graph_construct(sheet_num)[0]
    print(rdf_data, 'hi')

    response = HttpResponse(rdf_data, content_type='text/turtle')
    response['Content-Disposition'] = 'attachment; filename="rdf_graph_file.ttl"'

    return response
#rdflib simple example ผมสนใจ Monalisa พระพุทธรูปกับฟัน!!!!!!!!!! ระบบตรวจจับรอยโรคในฟันนนนน, ระบบวิเคราะห์ฉากทัศน์องค์ประกอบพระพุทธรูป กะเพราะปลา

#ต้องมีหน้า import โดยเฉพาะ ตัดง่ายๆไปเลย
@login_required(login_url="/sign-in")
def import_xlsx(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['excel_file']
            if uploaded_file.name.endswith('.xlsx'):
                excel_file = ExcelFile(name=uploaded_file.name, file=uploaded_file)
                excel_file.save()
                return redirect('import_xlsx')
            else:
                excel_files = ExcelFile.objects.all()
                return render(request, 'import.html', {'form': form, 'excel_files': excel_files, 'error_message': 'Please upload an Excel file (.xlsx)'})
    else:
        form = UploadFileForm()
    excel_files = ExcelFile.objects.all()
    return render(request, 'import.html', {'form': form, 'excel_files': excel_files})

def xlsx_del(request, file_number):
    try:
        # Convert file_number to an integer
        file_number = int(file_number)
        
        # Fetch the ExcelFile object corresponding to the file_number
        excel_file = ExcelFile.objects.get(pk=file_number)
        
        # Delete the ExcelFile object from the database
        os.remove(excel_file.file.path)
        excel_file.delete()
        
        # Return a redirect response to the import_xlsx view
        return redirect('import_xlsx')
    except ExcelFile.DoesNotExist:
        # If the ExcelFile object does not exist, return an error response
        return JsonResponse({'success': False, 'error': 'Excel file does not exist.'})
    except ValueError:
        # If file_number cannot be converted to an integer, return an error response
        return JsonResponse({'success': False, 'error': 'Invalid file number.'})
    
def excel_equip(request, file_number):
    file_number = int(file_number)
    return redirect(reverse('homepage', kwargs={'file_number': file_number}))
    
def visualization(request, sheet_num):
    # Load RDF graph from file
    g = graph_construct(sheet_num)[1]

    # Create nodes and edges for the network graph
    nx_graph = nx.Graph()

    # Add nodes and edges to the NetworkX graph
    for subj, pred, obj in g:
        subj = subj.split('#')[-1]
        obj = obj.split('#')[-1]
        pred = pred.split('#')[-1]
        nx_graph.add_node(subj)
        nx_graph.add_node(obj)
        nx_graph.add_edge(subj, obj, label=pred)

    pos = nx.spring_layout(nx_graph, k=0.5)

    fig = go.Figure()

    # Add nodes to the figure
    for node in nx_graph.nodes:
        x, y = pos[node]
        fig.add_trace(go.Scatter(x=[x], y=[y],
                             mode="markers+text", marker=dict(size=15),
                             text=node, textposition="bottom center",
                             textfont=dict(size=12)))  # Adjust the font size here

    # Add edges to the figure
    for edge in nx_graph.edges:
        x0, y0 = pos[edge[0]]
        x1, y1 = pos[edge[1]]
        fig.add_trace(go.Scatter(x=[x0, x1, None], y=[y0, y1, None],
                                line=dict(width=0.5, color='#888'), mode='lines'))

        # Add edge label
        edge_label = nx_graph.get_edge_data(edge[0], edge[1])['label']
        fig.add_trace(go.Scatter(x=[(x0 + x1) / 2], y=[(y0 + y1) / 2],
                                mode="text", text=[edge_label],
                                textposition="bottom center"))

    # Set layout options
    fig.update_layout(showlegend=False, hovermode='closest',margin=dict(l=0, r=0, b=0, t=0))

    plot_div = fig.to_html(full_html=False, include_plotlyjs=False)

    return render(request, 'visualization.html', {'plot_div': plot_div, 'sheet_num': sheet_num})