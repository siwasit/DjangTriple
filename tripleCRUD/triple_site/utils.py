import datetime
from .models import ExcelFile
from openpyxl import load_workbook
from rdflib import Graph, Literal, Namespace, RDF, XSD
import rdflib
from io import BytesIO

def parse_date(date_str):
    formats = ["%d %B %Y", "%Y-%m-%d", "%d %b %Y"]  # Add more formats as needed
    for fmt in formats:
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            pass
    return None

def graph_construct(sheet_num):
    excel_file = ExcelFile.objects.all()[sheet_num]
    excel_file_path = excel_file.file.path
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    graph = Graph()
    ex_person = Namespace("http://example.org/person#")
    foaf = Namespace("http://xmlns.com/foaf/0.1/")

    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i==1:
            continue

        name, class_type = row[0].split(':')
        predicate, obj = row[1], row[2]
        person_to_check = (ex_person[name], RDF.type, ex_person[class_type])

        triple_exists = False
        for triple in graph.triples((ex_person[name], RDF.type, None)):
            if triple == person_to_check:
                triple_exists = True
                break

        if not triple_exists:
            # If the triple does not exist, add it to the graph
            graph.add(person_to_check)

        if (ex_person[obj], RDF.type, None) in graph:
            graph.add((ex_person[name], foaf[predicate], ex_person[obj]))
        else:
            if parse_date(obj) is not None:
                graph.add((ex_person[name], ex_person[predicate], Literal(obj, datatype=XSD.date)))
            else:
                graph.add((ex_person[name], ex_person[predicate], Literal(obj, datatype=XSD.string)))

    rdf_data = BytesIO()
    graph.serialize(destination=rdf_data, format='turtle')
    
    # Prepare the HTTP response with the serialized RDF data

    return rdf_data.getvalue(), graph